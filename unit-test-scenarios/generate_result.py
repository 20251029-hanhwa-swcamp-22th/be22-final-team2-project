"""
단위/통합 테스트 결과서 분리 생성기.

출력:
- 단위_테스트_결과서.xlsx : 단위 테스트 실측 결과만 (요약 + 서비스별 4시트)
- 통합_테스트_결과서.xlsx : 통합 테스트 실측 결과만 (요약 + 서비스별 4시트)

분류 기준:
- 클래스 FQN에 ".integration." 들어가거나
- 클래스 simpleName에 "Integration" 들어가면
  → 통합 테스트
- 그 외 → 단위 테스트
"""
from __future__ import annotations
import datetime as dt
import re
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("D:/Users/Documents/be22-final-team2-project")
OUT = Path(__file__).parent

SERVICES = [
    ("Auth", "team2-backend-auth"),
    ("Activity", "team2-backend-activity"),
    ("Documents", "team2-backend-documents"),
    ("Master", "team2-backend-master"),
]

# 이 베이스 클래스를 extends 하는 테스트는 통합으로 분류
INTEGRATION_BASE_CLASSES = {
    "IntegrationTestSupport",
    "IntegrationTestBase",
    "AbstractIntegrationTest",
}


def scan_test_sources(service_dir: Path) -> dict[str, str]:
    """src/test/java 스캔 → {fqn: 'integration'|'unit'} 매핑.

    분류 우선순위:
    1) 패키지에 '.integration.' 포함 → integration
    2) @SpringBootTest 어노테이션 직접 사용 → integration
    3) IntegrationTestSupport 등 known base 상속 → integration
    4) 그 외 → unit
    """
    test_root = service_dir / "src" / "test" / "java"
    result: dict[str, str] = {}
    if not test_root.exists():
        return result

    pkg_pat = re.compile(r"^\s*package\s+([\w.]+);", re.MULTILINE)
    cls_pat = re.compile(r"\b(?:public\s+|abstract\s+|final\s+)*class\s+(\w+)")
    ext_pat = re.compile(r"\bclass\s+\w+\s+extends\s+(\w+)")

    for java_file in test_root.rglob("*.java"):
        try:
            text = java_file.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            try:
                text = java_file.read_text(encoding="cp949")
            except Exception:
                continue
        m_pkg = pkg_pat.search(text)
        m_cls = cls_pat.search(text)
        if not m_pkg or not m_cls:
            continue
        package = m_pkg.group(1)
        cls_name = m_cls.group(1)
        fqn = f"{package}.{cls_name}"

        is_int = False
        # 1) 패키지 기반
        if ".integration." in package or package.endswith(".integration"):
            is_int = True
        # 2) @SpringBootTest 어노테이션
        elif re.search(r"@SpringBootTest\b", text):
            is_int = True
        # 3) known integration base extends
        else:
            m_ext = ext_pat.search(text)
            if m_ext and m_ext.group(1) in INTEGRATION_BASE_CLASSES:
                is_int = True

        result[fqn] = "integration" if is_int else "unit"
    return result


def classify(cls_fqn: str, source_map: dict[str, str]) -> bool:
    """FQN을 source_map으로 분류. 없으면 fallback (이름 기반)."""
    cat = source_map.get(cls_fqn)
    if cat is not None:
        return cat == "integration"
    # fallback: 이름 기반
    if ".integration." in cls_fqn or cls_fqn.endswith(".integration"):
        return True
    simple = cls_fqn.rsplit(".", 1)[-1]
    return "Integration" in simple


def parse_junit(service_dir: Path, source_map: dict[str, str]) -> tuple[list[dict], dict]:
    """JUnit XML 디렉토리 → (테스트 메서드 리스트, 클래스별 요약)."""
    test_dir = service_dir / "build" / "test-results" / "test"
    methods: list[dict] = []
    by_class: dict[str, dict] = {}
    if not test_dir.exists():
        return methods, by_class
    for xml_file in sorted(test_dir.glob("TEST-*.xml")):
        try:
            tree = ET.parse(xml_file)
        except ET.ParseError:
            continue
        root = tree.getroot()
        # JUnit XML의 testsuite name은 @DisplayName으로 덮어쓰일 수 있으므로
        # 파일명(TEST-<FQN>.xml)에서 클래스 FQN을 추출한다.
        stem = xml_file.stem
        cls_name = stem[5:] if stem.startswith("TEST-") else stem
        tests = int(root.attrib.get("tests", "0"))
        failures = int(root.attrib.get("failures", "0"))
        errors = int(root.attrib.get("errors", "0"))
        skipped = int(root.attrib.get("skipped", "0"))
        time_s = float(root.attrib.get("time", "0"))
        timestamp = root.attrib.get("timestamp", "")
        is_int = classify(cls_name, source_map)
        by_class[cls_name] = {
            "tests": tests, "failures": failures, "errors": errors,
            "skipped": skipped, "time": time_s, "timestamp": timestamp,
            "is_integration": is_int,
        }
        for tc in root.findall("testcase"):
            mname = tc.attrib.get("name", "")
            mtime = float(tc.attrib.get("time", "0"))
            failure = tc.find("failure")
            error = tc.find("error")
            skipped_el = tc.find("skipped")
            if failure is not None:
                status = "FAIL"
                reason = (failure.attrib.get("message") or failure.text or "").strip().splitlines()[0][:300] if (failure.attrib.get("message") or failure.text) else ""
            elif error is not None:
                status = "ERROR"
                reason = (error.attrib.get("message") or error.text or "").strip().splitlines()[0][:300] if (error.attrib.get("message") or error.text) else ""
            elif skipped_el is not None:
                status = "SKIP"
                reason = ""
            else:
                status = "PASS"
                reason = ""
            methods.append({
                "class": cls_name,
                "method": mname,
                "status": status,
                "time": mtime,
                "reason": reason,
                "timestamp": timestamp,
                "is_integration": is_int,
            })
    return methods, by_class


def parse_jacoco(service_dir: Path) -> dict:
    xml_file = service_dir / "build" / "reports" / "jacoco" / "test" / "jacocoTestReport.xml"
    if not xml_file.exists():
        return {}
    try:
        # JaCoCo XML 의 DTD 참조를 무시하기 위해 entity resolution 비활성화
        parser = ET.XMLParser()
        tree = ET.parse(xml_file, parser=parser)
    except (ET.ParseError, OSError):
        return {}
    root = tree.getroot()
    result = {}
    for counter in root.findall("counter"):
        ctype = counter.attrib.get("type", "")
        missed = int(counter.attrib.get("missed", "0"))
        covered = int(counter.attrib.get("covered", "0"))
        total = missed + covered
        ratio = (covered / total * 100) if total else 0.0
        result[ctype] = {"missed": missed, "covered": covered, "total": total, "ratio": ratio}
    return result


# ============================================================
# 스타일
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")
NA_FILL = PatternFill("solid", fgColor="E7E6E6")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)
ALIGN_TL = Alignment(wrap_text=True, vertical="top", horizontal="left")
ALIGN_C = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_C
        cell.border = BORDER


def status_fill(status: str):
    if status == "PASS":
        return PASS_FILL
    if status in ("FAIL", "ERROR"):
        return FAIL_FILL
    if status == "SKIP":
        return SKIP_FILL
    return NA_FILL


def build_workbook(label: str, filter_integration: bool, svc_data: dict) -> openpyxl.Workbook:
    """label: '단위' or '통합'. filter_integration: True면 통합만, False면 단위만."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    now_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ---- 요약 시트 ----
    summary = wb.create_sheet("요약")
    summary.append([
        "서비스", "테스트 클래스 수", "테스트 메서드 수",
        "PASS", "FAIL", "ERROR", "SKIP",
        "총 수행시간(s)", "라인 커버리지 %", "분기 커버리지 %", "실행 시각",
    ])
    style_header(summary)

    grand_total = grand_pass = grand_fail = grand_err = grand_skip = 0
    grand_time = 0.0

    for svc_name, _svc_dir in SERVICES:
        data = svc_data[svc_name]
        methods = [m for m in data["methods"] if m["is_integration"] == filter_integration]
        classes = sorted({m["class"] for m in methods})
        total = len(methods)
        passed = sum(1 for m in methods if m["status"] == "PASS")
        failed = sum(1 for m in methods if m["status"] == "FAIL")
        errored = sum(1 for m in methods if m["status"] == "ERROR")
        skipped = sum(1 for m in methods if m["status"] == "SKIP")
        time_total = sum(m["time"] for m in methods)
        latest_ts = max((m.get("timestamp") or "" for m in methods), default="")

        line_cov = data["coverage"].get("LINE", {}).get("ratio", 0.0)
        branch_cov = data["coverage"].get("BRANCH", {}).get("ratio", 0.0)

        summary.append([
            svc_name,
            len(classes),
            total,
            passed,
            failed,
            errored,
            skipped,
            round(time_total, 2),
            round(line_cov, 2) if not filter_integration else "-",  # 커버리지는 단위에만 의미
            round(branch_cov, 2) if not filter_integration else "-",
            latest_ts or now_str,
        ])

        grand_total += total
        grand_pass += passed
        grand_fail += failed
        grand_err += errored
        grand_skip += skipped
        grand_time += time_total

    summary.append([
        "합계", "", grand_total, grand_pass, grand_fail, grand_err, grand_skip,
        round(grand_time, 2), "-", "-", now_str,
    ])
    for col, w in enumerate([12, 16, 18, 8, 8, 9, 8, 16, 18, 18, 22], start=1):
        summary.column_dimensions[get_column_letter(col)].width = w
    for r in summary.iter_rows(min_row=2, max_row=summary.max_row):
        for c in r:
            c.alignment = ALIGN_C
            c.border = BORDER
    for c in summary[summary.max_row]:
        c.font = Font(bold=True)
        c.fill = TOTAL_FILL

    # ---- 서비스별 실측 시트 ----
    for svc_name, _svc_dir in SERVICES:
        data = svc_data[svc_name]
        methods = [m for m in data["methods"] if m["is_integration"] == filter_integration]
        ws = wb.create_sheet(svc_name)
        ws.append(["테스트 클래스", "테스트 메서드", "결과", "수행시간(s)", "실패 사유", "실행 시각"])
        style_header(ws)
        for m in methods:
            ws.append([
                m["class"].rsplit(".", 1)[-1],
                m["method"],
                m["status"],
                round(m["time"], 3),
                m["reason"],
                m.get("timestamp", ""),
            ])
        widths = [44, 64, 10, 14, 60, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for ridx, r in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for c in r:
                c.alignment = ALIGN_TL
                c.border = BORDER
            sc = ws.cell(row=ridx, column=3)
            sc.fill = status_fill(sc.value if isinstance(sc.value, str) else "")
            sc.alignment = ALIGN_C

    return wb


def main():
    svc_data: dict[str, dict] = {}
    for svc_name, svc_dir in SERVICES:
        sdir = ROOT / svc_dir
        source_map = scan_test_sources(sdir)
        methods, by_class = parse_junit(sdir, source_map)
        cov = parse_jacoco(sdir)
        svc_data[svc_name] = {
            "methods": methods,
            "by_class": by_class,
            "coverage": cov,
            "source_map": source_map,
        }

    wb_unit = build_workbook("단위", filter_integration=False, svc_data=svc_data)
    wb_unit.save(OUT / "단위_테스트_결과서.xlsx")

    wb_int = build_workbook("통합", filter_integration=True, svc_data=svc_data)
    wb_int.save(OUT / "통합_테스트_결과서.xlsx")

    # 통계 출력 (검증용)
    print("=" * 60)
    print("결과서 생성 완료")
    print("=" * 60)
    for svc_name, _ in SERVICES:
        ms = svc_data[svc_name]["methods"]
        unit_ms = [m for m in ms if not m["is_integration"]]
        int_ms = [m for m in ms if m["is_integration"]]
        unit_pass = sum(1 for m in unit_ms if m["status"] == "PASS")
        unit_fail = sum(1 for m in unit_ms if m["status"] in ("FAIL", "ERROR"))
        int_pass = sum(1 for m in int_ms if m["status"] == "PASS")
        int_fail = sum(1 for m in int_ms if m["status"] in ("FAIL", "ERROR"))
        unit_classes = len({m["class"] for m in unit_ms})
        int_classes = len({m["class"] for m in int_ms})
        print(f"  {svc_name}:")
        print(f"    단위: {len(unit_ms)} methods / {unit_classes} classes  PASS={unit_pass} FAIL={unit_fail}")
        print(f"    통합: {len(int_ms)} methods / {int_classes} classes  PASS={int_pass} FAIL={int_fail}")
        print(f"    총합: {len(ms)} methods")


if __name__ == "__main__":
    main()
