"""
결과서 트리플 검증 스크립트.

검증 항목:
1. 단위_테스트_결과서.xlsx / 통합_테스트_결과서.xlsx 둘 다 존재
2. 각 결과서가 5개 시트 (요약 + 4개 서비스)만 가짐 (시나리오 시트 없음)
3. 단위 결과서의 4개 서비스 시트에 통합 클래스 없음
4. 통합 결과서의 4개 서비스 시트에 단위 클래스 없음
5. xlsx 시트의 row 수 == 원본 XML 메서드 수와 일치
6. xlsx의 모든 결과 = PASS (FAIL/ERROR/SKIP 0개)
7. 단위 + 통합 = 원본 총 메서드 수와 일치
"""
from __future__ import annotations
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path("D:/Users/Documents/be22-final-team2-project")
OUT = Path(__file__).parent

SERVICES = [
    ("Auth", "team2-backend-auth"),
    ("Activity", "team2-backend-activity"),
    ("Documents", "team2-backend-documents"),
    ("Master", "team2-backend-master"),
]

INTEGRATION_BASE_CLASSES = {
    "IntegrationTestSupport",
    "IntegrationTestBase",
    "AbstractIntegrationTest",
}


def scan_test_sources(service_dir: Path) -> dict[str, str]:
    """src/test/java 스캔 → {fqn: 'integration'|'unit'}."""
    test_root = service_dir / "src" / "test" / "java"
    result: dict[str, str] = {}
    if not test_root.exists():
        return result
    pkg_pat = re.compile(r"^\s*package\s+([\w.]+);", re.MULTILINE)
    cls_pat = re.compile(r"\b(?:public\s+|abstract\s+|final\s+)*class\s+(\w+)")
    ext_pat = re.compile(r"\bclass\s+\w+\s+extends\s+(\w+)")
    for java_file in test_root.rglob("*.java"):
        try:
            text = java_file.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            try:
                text = java_file.read_text(encoding="cp949")
            except Exception:
                continue
        m_pkg = pkg_pat.search(text)
        m_cls = cls_pat.search(text)
        if not m_pkg or not m_cls:
            continue
        package = m_pkg.group(1)
        cls_name = m_cls.group(1)
        fqn = f"{package}.{cls_name}"
        is_int = False
        if ".integration." in package or package.endswith(".integration"):
            is_int = True
        elif re.search(r"@SpringBootTest\b", text):
            is_int = True
        else:
            m_ext = ext_pat.search(text)
            if m_ext and m_ext.group(1) in INTEGRATION_BASE_CLASSES:
                is_int = True
        result[fqn] = "integration" if is_int else "unit"
    return result


def classify(cls_fqn: str, source_map: dict[str, str]) -> bool:
    cat = source_map.get(cls_fqn)
    if cat is not None:
        return cat == "integration"
    if ".integration." in cls_fqn or cls_fqn.endswith(".integration"):
        return True
    simple = cls_fqn.rsplit(".", 1)[-1]
    return "Integration" in simple


def parse_junit_truth(service_dir: Path, source_map: dict[str, str]):
    """원본 XML에서 직접 진실을 추출."""
    test_dir = service_dir / "build" / "test-results" / "test"
    truth = {"unit": [], "integration": []}
    for xml_file in sorted(test_dir.glob("TEST-*.xml")):
        try:
            tree = ET.parse(xml_file)
        except ET.ParseError:
            continue
        root = tree.getroot()
        stem = xml_file.stem
        cls_name = stem[5:] if stem.startswith("TEST-") else stem
        category = "integration" if classify(cls_name, source_map) else "unit"
        for tc in root.findall("testcase"):
            mname = tc.attrib.get("name", "")
            failure = tc.find("failure") is not None
            error = tc.find("error") is not None
            skipped = tc.find("skipped") is not None
            if failure:
                status = "FAIL"
            elif error:
                status = "ERROR"
            elif skipped:
                status = "SKIP"
            else:
                status = "PASS"
            truth[category].append({
                "class_simple": cls_name.rsplit(".", 1)[-1],
                "class_fqn": cls_name,
                "method": mname,
                "status": status,
            })
    return truth


def load_sheet(wb, name):
    ws = wb[name]
    rows = []
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return headers, rows


def main():
    print("=" * 70)
    print("결과서 검증 시작")
    print("=" * 70)

    # 1. 파일 존재 확인
    unit_path = OUT / "단위_테스트_결과서.xlsx"
    int_path = OUT / "통합_테스트_결과서.xlsx"
    assert unit_path.exists(), f"❌ {unit_path} not found"
    assert int_path.exists(), f"❌ {int_path} not found"
    print(f"✅ 결과서 2개 파일 존재")

    wb_unit = openpyxl.load_workbook(unit_path)
    wb_int = openpyxl.load_workbook(int_path)

    # 2. 시트 구성 확인 (요약 + 4개 서비스, 총 5개)
    expected_sheets = ["요약", "Auth", "Activity", "Documents", "Master"]
    assert wb_unit.sheetnames == expected_sheets, f"❌ 단위 결과서 시트: {wb_unit.sheetnames}"
    assert wb_int.sheetnames == expected_sheets, f"❌ 통합 결과서 시트: {wb_int.sheetnames}"
    print(f"✅ 두 결과서 모두 5개 시트 (요약 + 4개 서비스). 시나리오 시트 없음.")

    # 3. 원본 진실 수집 (어노테이션 기반 분류)
    truth = {}
    for svc_name, svc_dir in SERVICES:
        sm = scan_test_sources(ROOT / svc_dir)
        truth[svc_name] = parse_junit_truth(ROOT / svc_dir, sm)

    errors = []
    warnings = []

    # 4. 단위 결과서 검증
    print("\n--- 단위 결과서 검증 ---")
    for svc_name, _ in SERVICES:
        _, rows = load_sheet(wb_unit, svc_name)
        truth_unit = truth[svc_name]["unit"]
        truth_int = truth[svc_name]["integration"]

        # 4a. row 수 일치
        if len(rows) != len(truth_unit):
            errors.append(f"❌ 단위/{svc_name}: row {len(rows)} ≠ truth {len(truth_unit)}")
        else:
            print(f"  ✅ 단위/{svc_name}: {len(rows)} rows")

        # 4b. 통합 클래스가 섞여있지 않은지 (각 row의 클래스가 truth_unit에 있는지)
        truth_unit_classes = {m["class_simple"] for m in truth_unit}
        truth_int_classes = {m["class_simple"] for m in truth_int}
        for row in rows:
            cls = row["테스트 클래스"]
            if cls in truth_int_classes:
                errors.append(f"❌ 단위/{svc_name}: 통합 클래스 '{cls}' 섞임")
            elif cls not in truth_unit_classes:
                warnings.append(f"⚠ 단위/{svc_name}: 알 수 없는 클래스 '{cls}'")

        # 4c. 모든 결과 PASS
        non_pass = [r for r in rows if r["결과"] != "PASS"]
        if non_pass:
            errors.append(f"❌ 단위/{svc_name}: PASS 아닌 항목 {len(non_pass)}개: {[r['테스트 메서드'] for r in non_pass[:3]]}")

    # 5. 통합 결과서 검증
    print("\n--- 통합 결과서 검증 ---")
    for svc_name, _ in SERVICES:
        _, rows = load_sheet(wb_int, svc_name)
        truth_int = truth[svc_name]["integration"]
        truth_unit = truth[svc_name]["unit"]

        # 5a. row 수 일치
        if len(rows) != len(truth_int):
            errors.append(f"❌ 통합/{svc_name}: row {len(rows)} ≠ truth {len(truth_int)}")
        else:
            print(f"  ✅ 통합/{svc_name}: {len(rows)} rows")

        # 5b. 단위 클래스가 섞여있지 않은지
        truth_int_classes = {m["class_simple"] for m in truth_int}
        truth_unit_classes = {m["class_simple"] for m in truth_unit}
        for row in rows:
            cls = row["테스트 클래스"]
            if cls in truth_unit_classes:
                errors.append(f"❌ 통합/{svc_name}: 단위 클래스 '{cls}' 섞임")
            elif cls not in truth_int_classes:
                warnings.append(f"⚠ 통합/{svc_name}: 알 수 없는 클래스 '{cls}'")

        # 5c. 모든 결과 PASS
        non_pass = [r for r in rows if r["결과"] != "PASS"]
        if non_pass:
            errors.append(f"❌ 통합/{svc_name}: PASS 아닌 항목 {len(non_pass)}개: {[r['테스트 메서드'] for r in non_pass[:3]]}")

    # 6. 합산 검증 (단위 + 통합 = 원본)
    print("\n--- 합산 검증 ---")
    for svc_name, _ in SERVICES:
        _, unit_rows = load_sheet(wb_unit, svc_name)
        _, int_rows = load_sheet(wb_int, svc_name)
        total = len(unit_rows) + len(int_rows)
        truth_total = len(truth[svc_name]["unit"]) + len(truth[svc_name]["integration"])
        if total != truth_total:
            errors.append(f"❌ {svc_name} 합산: {total} ≠ truth {truth_total}")
        else:
            print(f"  ✅ {svc_name}: 단위 {len(unit_rows)} + 통합 {len(int_rows)} = {total} (원본 {truth_total})")

    # 7. 요약 시트 검증
    print("\n--- 요약 시트 검증 ---")
    for label, wb, key in [("단위", wb_unit, "unit"), ("통합", wb_int, "integration")]:
        _, rows = load_sheet(wb, "요약")
        # 마지막 row는 합계
        body_rows = rows[:-1]
        total_row = rows[-1]
        for r, (svc_name, _) in zip(body_rows, SERVICES):
            if r["서비스"] != svc_name:
                errors.append(f"❌ 요약 {label}: 서비스명 mismatch {r['서비스']} ≠ {svc_name}")
                continue
            t_count = len(truth[svc_name][key])
            if r["테스트 메서드 수"] != t_count:
                errors.append(f"❌ 요약 {label}/{svc_name}: 메서드 수 {r['테스트 메서드 수']} ≠ truth {t_count}")
            t_pass = sum(1 for m in truth[svc_name][key] if m["status"] == "PASS")
            if r["PASS"] != t_pass:
                errors.append(f"❌ 요약 {label}/{svc_name}: PASS {r['PASS']} ≠ truth {t_pass}")
        # 합계 row
        grand = sum(len(truth[s[0]][key]) for s in SERVICES)
        if total_row["테스트 메서드 수"] != grand:
            errors.append(f"❌ 요약 {label}/합계: {total_row['테스트 메서드 수']} ≠ truth {grand}")
        else:
            print(f"  ✅ 요약 {label}/합계: {grand} 메서드")

    # 8. 결과 출력
    print("\n" + "=" * 70)
    if errors:
        print(f"❌ 검증 실패 — {len(errors)}건 오류")
        for e in errors:
            print(f"  {e}")
    else:
        print("✅ 모든 검증 통과 (트리플 체크 OK)")
    if warnings:
        print(f"⚠ 경고 {len(warnings)}건")
        for w in warnings[:10]:
            print(f"  {w}")

    return 0 if not errors else 1


if __name__ == "__main__":
    sys.exit(main())
